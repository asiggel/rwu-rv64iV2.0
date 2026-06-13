#!/usr/bin/env python3
"""Generate RWU-RV64I Cache / AMAT / Clock-Matrix Excel workbook."""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "RWU_RV64I_Cache_Analysis.xlsx")

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def fill(hex6): return PatternFill("solid", fgColor=hex6)
def fwh(bold=True, sz=10): return Font(color="FFFFFF", bold=bold, size=sz)
def fbk(bold=False, sz=10, color="000000"): return Font(color=color, bold=bold, size=sz)

FILL_DARK   = fill("1F4E79")   # dark blue  – main header
FILL_MED    = fill("2E75B6")   # mid  blue  – sub-header
FILL_LIGHT  = fill("D6E4F0")   # light blue – section bg
FILL_GREEN  = fill("E2EFDA")   # light green
FILL_DGRN   = fill("375623")   # dark green – best option
FILL_BGRN   = fill("70AD47")   # bright green – good
FILL_YELLOW = fill("FFFF99")   # yellow – marginal
FILL_ORANGE = fill("FFC000")   # orange – warning
FILL_RED    = fill("FFB3B3")   # red – bad
FILL_GREY   = fill("F2F2F2")   # alternating row
FILL_WHITE  = fill("FFFFFF")

thn = Side(border_style="thin",   color="000000")
med = Side(border_style="medium", color="000000")
THIN = Border(left=thn, right=thn, top=thn, bottom=thn)
MED  = Border(left=med, right=med, top=med, bottom=med)
LMED = Border(left=med, right=thn, top=thn, bottom=thn)

def s(ws, row, col, val=None, f=None, fnt=None, ha="left", va="center",
      brd=THIN, nf=None, wrap=False, rs=1, cs=1):
    if rs > 1 or cs > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row+rs-1, end_column=col+cs-1)
    c = ws.cell(row=row, column=col)
    if val is not None: c.value = val
    if f:   c.fill  = f
    if fnt: c.font  = fnt
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if brd: c.border = brd
    if nf:  c.number_format = nf
    return c

def hdr(ws, row, col, val, cs=1):
    return s(ws, row, col, val, FILL_DARK, fwh(sz=11), ha="center", cs=cs)

def shdr(ws, row, col, val, cs=1):
    return s(ws, row, col, val, FILL_MED, fwh(sz=10), ha="center", cs=cs)

def sh2(ws, row, col, val, cs=1):
    return s(ws, row, col, val, FILL_LIGHT, fbk(bold=True), cs=cs)

def lbl(ws, row, col, val, bold=True, wrap=False):
    return s(ws, row, col, val, fnt=fbk(bold=bold), wrap=wrap)

def num(ws, row, col, val, nf="0.00", f=None):
    return s(ws, row, col, val, f=f, fnt=fbk(), ha="center", nf=nf)

# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()

# ===========================================================================
# SHEET 1 – Legende
# ===========================================================================
ws1 = wb.active
ws1.title = "Legende"
for col, w in zip("ABCDE", [30, 28, 48, 22, 22]):
    ws1.column_dimensions[get_column_letter(col.encode()[0]-64)].width = w

r = 1
# Title
s(ws1, r, 1, "RWU-RV64I  —  Cache-Speichersystem  (Hennessy-Patterson Analyse)",
  FILL_DARK, fwh(sz=13), ha="center", cs=5); ws1.row_dimensions[r].height=28; r+=1
s(ws1, r, 1, "Prozess: X-Fab XO035 (350 nm)  |  Architektur: Harvard  |  "
  "Konfiguration: I-Cache + D-Cache (je 4-way SA, 4 KB), Scratchpad SRAM",
  FILL_MED, fwh(sz=10, bold=False), ha="center", cs=5); r+=2

# --- Cache-Eigenschaften -------------------------------------------------
shdr(ws1, r, 1, "Eigenschaft", cs=1)
shdr(ws1, r, 2, "I-Cache", cs=1)
shdr(ws1, r, 3, "D-Cache (Flash-Region)", cs=1)
shdr(ws1, r, 4, "D-Cache (Scratchpad-Region)", cs=1)
shdr(ws1, r, 5, "Scratchpad SRAM (Bypass)", cs=1)
r+=1
rows_cache = [
    ("Kapazität",              "4 KB (param.)",           "4 KB (param.)",           "n/a (bypass)",              "TBD (param.)"),
    ("Assoziativität",         "4-fach set-assoziativ",   "4-fach set-assoziativ",   "—",                         "direkt adressiert"),
    ("Sets / Cacheline",       "32 Sets, 32 B",           "32 Sets, 32 B",           "—",                         "—"),
    ("Ersetzungsstrategie",    "Pseudo-LRU (3 Bit/Set)",  "Pseudo-LRU (3 Bit/Set)",  "—",                         "—"),
    ("Write Policy",           "read-only (kein Write)",  "read-allocate, no-write-back",  "kein Cache (bypass)",  "write-through direkt"),
    ("Write-Allocate",         "nicht anwendbar",         "nein  (Flash read-only)",  "nein (bypass)",            "nicht anwendbar"),
    ("Dirty Bit",              "nein",                    "nein (eliminiert)",        "nein",                     "—"),
    ("Hit-Latenz",             "1 Systemtakt",            "1 Systemtakt",            "1–2 Systemtakte (bypass)",  "1–2 Systemtakte"),
    ("Miss-Penalty",           "106 × f_sys/f_QSPI Takte","106 × f_sys/f_QSPI Takte","—  (kein Miss möglich)",   "—"),
    ("AXI4 Write-Path",        "nein",                    "nein (eliminiert)",        "nein",                     "nein"),
    ("AXI4 Read-Path",         "ja  (AR + R Kanal)",      "ja  (AR + R Kanal)",       "nein",                     "nein"),
    ("Bus zum Speicher",       "AXI4 Burst (ARLEN=3)",    "AXI4 Burst (ARLEN=3)",    "SRAM-Interface (sync.)",    "SRAM-Interface (sync.)"),
]
for prop, ic, dc_fl, dc_sp, sp in rows_cache:
    bg = FILL_GREY if (r % 2 == 0) else FILL_WHITE
    lbl(ws1, r, 1, prop, bold=True)
    for ci, val in enumerate([ic, dc_fl, dc_sp, sp], start=2):
        s(ws1, r, ci, val, f=bg, fnt=fbk(bold=False), wrap=True)
    r+=1
r+=1

# --- H&P Terminologie -------------------------------------------------------
shdr(ws1, r, 1, "H&P Begriff", cs=2)
shdr(ws1, r, 3, "Bedeutung / Formel", cs=2)
shdr(ws1, r, 5, "In diesem System")
r+=1
terms = [
    ("Hit Time  (t_hit)",
     "Zugriffszeit bei Cache-Treffer",
     "1 Systemtakt (I+D)"),
    ("Miss Rate  (m)",
     "Anteil Zugriffe ohne Treffer",
     "I-Cache: abhängig vom Code-Footprint\nD-Cache (Flash): abhängig von .rodata-Zugriffen"),
    ("Miss Penalty  (t_miss)",
     "Mehrzyklen bei Miss bis Daten bereit stehen",
     "106 × (f_sys / f_QSPI_SCK)  Takte\n≡ 4.24 µs bei 25 MHz QSPI (beide Flash-Typen)"),
    ("AMAT",
     "Average Memory Access Time\n= t_hit + m × t_miss",
     "I: 1 + m_I × t_miss\nD: 1 + m_D × t_miss  (nur Load-Misses aus Flash)"),
    ("Memory Stall CPI",
     "Taktzyklen/Instr. für Speicherwartezustände\n= m_I×t_miss + load_freq×m_D×t_miss",
     "Store-Misses = 0 (Scratchpad-Bypass, kein Flash-Write)"),
    ("CPI_eff",
     "Effektiver CPI\n= CPI_ideal + Memory_Stall_CPI",
     "CPI_ideal = 1 (No-Pipeline-CPU)"),
    ("Write-Through",
     "Jeder Write sofort in Backing Store",
     "NICHT verwendet (kein Flash-Write)"),
    ("Write-Back",
     "Write erst bei Eviction in Backing Store",
     "NICHT verwendet (Flash read-only, Dirty-Bit eliminiert)"),
    ("Write-Allocate",
     "Miss bei Write → Cache-Line laden",
     "NICHT verwendet (Stores gehen direkt → Scratchpad)"),
    ("No-Write-Allocate",
     "Miss bei Write → direkt in Backing Store",
     "NICHT relevant (Scratchpad ist nicht Flash)"),
    ("Read-Allocate",
     "Miss bei Load → neue Line in Cache laden",
     "JA – für I-Cache und D-Cache (Flash-Region)"),
    ("Pseudo-LRU (PLRU)",
     "3-Bit-Approximation des LRU für 4-Wege",
     "1 PLRU-Feld pro Set (96 Bit Tag-SRAM enthält 3 PLRU-Bits)"),
    ("Scratchpad",
     "Direkt adressierter SRAM (kein Cache)",
     "Hält Stack, Heap, .data, .bss\nD-Cache-Controller bypassed direkt"),
]
for i, (term, defn, sys_note) in enumerate(terms):
    bg = FILL_GREY if (i % 2 == 0) else FILL_WHITE
    s(ws1, r, 1, term, f=FILL_LIGHT, fnt=fbk(bold=True), cs=2, wrap=True)
    s(ws1, r, 3, defn, f=bg, fnt=fbk(), cs=2, wrap=True)
    s(ws1, r, 5, sys_note, f=bg, fnt=fbk(), wrap=True)
    ws1.row_dimensions[r].height = 32
    r+=1
r+=1

# --- Linker-Segmente --------------------------------------------------------
shdr(ws1, r, 1, "Linker-Segment", cs=1)
shdr(ws1, r, 2, "Read/Write", cs=1)
shdr(ws1, r, 3, "Inhalt", cs=2)
shdr(ws1, r, 5, "Speicherort")
r+=1
segs = [
    (".text",    "RO", "Programmcode (Instruktionen)",                  "NOR Flash → I-Cache"),
    (".rodata",  "RO", "Konstanten, String-Literale",                   "NOR Flash → D-Cache (Flash-Region)"),
    (".data",    "RW", "Initialisierte globale/statische Variablen",    "Scratchpad (Kopie von Flash bei Boot)"),
    (".bss",     "RW", "Uninitial. glob./stat. Var. (Null-Init.)",      "Scratchpad (Start-up nullt Region)"),
    ("Heap",     "RW", "Dynamisch alloz. Speicher (malloc/free)",       "Scratchpad (wächst ↑ von .bss)"),
    ("Stack",    "RW", "Auto-Variablen, Rücksprungadr., Register-Save", "Scratchpad (wächst ↓ von oben)"),
]
for i, (seg, rw, content, where) in enumerate(segs):
    bg = FILL_GREEN if rw == "RO" else FILL_YELLOW
    s(ws1, r, 1, seg,     f=bg,       fnt=fbk(bold=True), ha="center")
    s(ws1, r, 2, rw,      f=bg,       fnt=fbk(bold=False), ha="center")
    s(ws1, r, 3, content, f=FILL_WHITE, fnt=fbk(), cs=2, wrap=True)
    s(ws1, r, 5, where,   f=FILL_WHITE, fnt=fbk())
    r+=1

# ===========================================================================
# SHEET 2 – AMAT Szenarien (Hennessy-Patterson)
# ===========================================================================
ws2 = wb.create_sheet("AMAT_Szenarien")
for col, w in zip(range(1, 14), [6, 22, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]):
    ws2.column_dimensions[get_column_letter(col)].width = w

r = 1
s(ws2, r, 1, "AMAT-Analyse nach Hennessy-Patterson  —  RWU-RV64I Cache-System",
  FILL_DARK, fwh(sz=12), ha="center", cs=13)
ws2.row_dimensions[r].height = 24; r+=2

# --- Parameter Box ----------------------------------------------------------
shdr(ws2, r, 1, "System-Parameter", cs=5)
r+=1
params = [
    ("t_hit (I-Cache + D-Cache)",    "1",   "Systemtakt",   "Kombinatorisch: Tag-Vergleich + SRAM-Read in 1 Takt"),
    ("t_hit (Scratchpad Bypass)",    "1–2", "Systemtakte",  "D-Cache-Controller → SRAM direkt"),
    ("QSPI-Protokoll-Zyklen (tot.)", "106", "QSPI-SCK-Takte","8 (Opcode)+24 (Adr)+8 (Dummy)+64 (Daten, Quad)+2 (CS#)"),
    ("Miss-Penalty  t_miss",         "106 × (f_sys / f_QSPI)", "Systemtakte", "Skalar von Taktver­hältnis, siehe Takt-Matrix"),
    ("Beispiel: f_sys=25 MHz, f_QSPI=25 MHz", "106", "Systemtakte", "≡ 4.24 µs"),
    ("Beispiel: f_sys=10 MHz, f_QSPI=25 MHz", " 42", "Systemtakte", "≡ 4.24 µs (Spec-Referenzpunkt)"),
    ("CPI_ideal  (No-Pipeline)",     "1",   "Takte/Instr.", "Ein Fetch pro Takt (ohne Speicherstalls)"),
    ("Instruction Mix (typisch)",    "—",   "—",            "25 % Load, 10 % Store, 20 % Branch, 45 % ALU"),
    ("Store-Miss-Penalty",           "0",   "—",            "Stores → Scratchpad Bypass (kein Flash-Write)"),
]
for i, (name, val, unit, note) in enumerate(params):
    bg = FILL_GREY if i % 2 == 0 else FILL_WHITE
    s(ws2, r, 1, name,  f=FILL_LIGHT, fnt=fbk(bold=True), cs=2, wrap=True)
    s(ws2, r, 3, val,   f=bg, fnt=fbk(bold=True), ha="center")
    s(ws2, r, 4, unit,  f=bg, fnt=fbk())
    s(ws2, r, 5, note,  f=bg, fnt=fbk(), wrap=True)
    r+=1
r+=1

# --- AMAT-Formel Box --------------------------------------------------------
shdr(ws2, r, 1, "Hennessy-Patterson Formeln (angewendet auf RWU-RV64I)", cs=13); r+=1
formeln = [
    "AMAT  =  t_hit  +  m  ×  t_miss",
    "CPI_eff  =  CPI_ideal  +  I$_stall  +  D$_stall",
    "I$_stall  =  m_I  ×  t_miss                          [Takte pro Instruktion]",
    "D$_stall  =  load_freq × m_D × t_miss                [Store-Miss = 0, Scratchpad-Bypass!]",
    "Memory_Stall_CPI  =  m_I × t_miss  +  0.25 × m_D × t_miss",
]
for f_str in formeln:
    s(ws2, r, 1, f_str, FILL_LIGHT, fbk(bold=True, sz=10), cs=13)
    ws2.row_dimensions[r].height = 18; r+=1
r+=1

# --- AMAT-Tabellen pro t_miss -----------------------------------------------
MISS_PENALTIES = [
    (42,  "f_sys=10 MHz, f_QSPI=25 MHz  (Spec-Referenz)"),
    (106, "f_sys=25 MHz, f_QSPI=25 MHz  (Empfehlung: kein PLL)"),
    (212, "f_sys=50 MHz, f_QSPI=25 MHz  (Empfehlung: PLL×2)"),
    (106, "f_sys=50 MHz, f_QSPI=50 MHz  (aggressiv, div=1)"),
]
MISS_RATES = [0.001, 0.005, 0.01, 0.02, 0.05, 0.10, 0.15, 0.20, 0.30, 0.50]
COL_LABELS = ["Miss-Rate m", "AMAT_I [Takte]", "AMAT_D [Takte]",
              "I$_stall/Instr", "D$_stall/Instr", "Mem_Stall_CPI",
              "CPI_eff", "IPC_eff", "Speed-Up vs. ideal", "Zeitverlust [%]"]

for t_miss, scenario_label in MISS_PENALTIES:
    # Scenario header
    s(ws2, r, 1, f"Szenario: {scenario_label}  |  t_miss = {t_miss} Systemtakte",
      FILL_DARK, fwh(sz=11), ha="left", cs=13)
    ws2.row_dimensions[r].height = 20; r+=1

    # Column headers
    for ci, lbl_txt in enumerate(COL_LABELS, start=1):
        shdr(ws2, r, ci, lbl_txt)
    shdr(ws2, r, 11, "CPI-Diagramm (normiert)")
    shdr(ws2, r, 12, "Klassifikation")
    ws2.row_dimensions[r].height = 30; r+=1

    for i, m in enumerate(MISS_RATES):
        amat_i   = 1 + m * t_miss
        amat_d   = 1 + m * t_miss
        i_stall  = m * t_miss
        d_stall  = 0.25 * m * t_miss        # 25 % load frequency
        mem_stall = i_stall + d_stall
        cpi_eff  = 1 + mem_stall
        ipc      = 1.0 / cpi_eff
        speedup  = cpi_eff                  # vs ideal CPI=1, SpeedUp = CPI_eff
        loss_pct = (cpi_eff - 1) * 100

        if m <= 0.01:    bg = FILL_GREEN
        elif m <= 0.05:  bg = FILL_BGRN
        elif m <= 0.10:  bg = FILL_YELLOW
        elif m <= 0.20:  bg = FILL_ORANGE
        else:            bg = FILL_RED

        cls = ("sehr gut"     if m <= 0.01 else
               "gut"          if m <= 0.05 else
               "akzeptabel"   if m <= 0.10 else
               "problematisch"if m <= 0.20 else
               "kritisch")

        vals = [f"{m*100:.1f} %", amat_i, amat_d, i_stall, d_stall,
                mem_stall, cpi_eff, ipc, speedup, loss_pct]
        nfmts = ["@", "0.00", "0.00", "0.00", "0.00",
                 "0.00", "0.00", "0.000", "0.00", "0.0"]
        for ci, (v, nf_) in enumerate(zip(vals, nfmts), start=1):
            num(ws2, r, ci, v, nf=nf_, f=bg)
        # Mini-bar (text)
        bar_len = max(1, int(mem_stall))
        bar = "█" * min(bar_len, 20) + ("+" if bar_len > 20 else "")
        s(ws2, r, 11, bar, f=bg, fnt=Font(color="1F4E79"))
        s(ws2, r, 12, cls, f=bg, fnt=fbk(bold=(m > 0.1)))
        r+=1

    # Legend for this block
    s(ws2, r, 1, "Grün ≤ 1 % | Hell-Grün ≤ 5 % | Gelb ≤ 10 % | Orange ≤ 20 % | Rot > 20 %",
      FILL_GREY, fbk(bold=False, sz=9), cs=13)
    r+=2

# ===========================================================================
# SHEET 3 – Takt-Matrix
# ===========================================================================
ws3 = wb.create_sheet("Takt_Matrix")
col_widths = [5, 18, 14, 14, 14, 18, 18, 18, 18, 14, 18, 18, 18, 26]
for ci, w in enumerate(col_widths, start=1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

r = 1
s(ws3, r, 1, "Takt-Kombinations-Matrix  —  RWU-RV64I  (QSPI / CPU-Core / SoC / Eingangs-Takt)",
  FILL_DARK, fwh(sz=12), ha="center", cs=14)
ws3.row_dimensions[r].height = 24; r+=2

# Flash device info box
shdr(ws3, r, 1, "Referenz-Flash-Geräte", cs=14); r+=1
flash_rows = [
    ("Winbond W25Q128JV",  "128 Mbit (16 MB)", "133 MHz",  "80 MHz (konservativ, 2.7–3.6 V)",
     "0x6B (Quad Out, 1-1-4)", "8", "0xEF", "★ Primär-Referenz laut Spec"),
    ("Micron MT25QL128",   "128 Mbit (16 MB)", "133 MHz",  "80 MHz (konservativ, 2.7–3.6 V)",
     "0x6B (Quad Out, 1-1-4)", "8", "0x20", "★ Sekundär-Referenz laut Spec"),
]
fhdr_cols = ["Gerät", "Kapazität", "Max SCK (Datenblatt)", "Max SCK (Design-Ziel)",
             "Quad-Read Opcode", "Dummy-Zyklen", "JEDEC-ID", "Anmerkung"]
for ci, h in enumerate(fhdr_cols, start=1):
    shdr(ws3, r, ci+1 if ci > 7 else ci,  # shift hack avoided below
         h if ci <= len(fhdr_cols) else "")
# rebuild properly:
ws3.row_dimensions[r].height = 0  # hide the broken row
r+=1
for ci, h in enumerate(fhdr_cols, start=1):
    shdr(ws3, r, ci, h)
r+=1
for i, fr in enumerate(flash_rows):
    bg = FILL_GREEN if i == 0 else FILL_LIGHT
    for ci, v in enumerate(fr, start=1):
        s(ws3, r, ci, v, f=bg, fnt=fbk(bold=(ci==1 or ci==8)), wrap=True)
    r+=1
r+=1

# --- Clock Matrix Headers ---------------------------------------------------
COL_HDRS = [
    "#", "Quarz f_in\n[MHz]", "PLL\n(ja/nein)", "PLL-Faktor\nN",
    "f_sys / f_CPU\n[MHz]", "QSPI-Takt\nQuelle", "QSPI-Divider\n(÷)",
    "f_QSPI_SCK\n[MHz]", "t_QSPI-Takt\n[ns]", "QSPI-Zyklen\n(Protokoll)",
    "t_miss\n[µs]", "Miss-Penalty\n[Systemtakte]",
    "W25Q128JV\n(≤80 MHz?)", "MT25QL128\n(≤80 MHz?)",
]
for ci, h in enumerate(COL_HDRS, start=1):
    c = shdr(ws3, r, ci, h)
    ws3.row_dimensions[r].height = 40
    ws3.cell(row=r, column=ci).alignment = Alignment(horizontal="center",
                                                      vertical="center",
                                                      wrap_text=True)
r+=1

# Second header row for AMAT at different miss rates + recommendation
AMAT_MR = [0.01, 0.05, 0.10, 0.20]
amat_start_col = len(COL_HDRS) + 1
for ci, mr in enumerate(AMAT_MR, start=amat_start_col):
    shdr(ws3, r-1, ci, f"AMAT\n@ {int(mr*100)} % Miss")
shdr(ws3, r-1, amat_start_col + len(AMAT_MR), "Timing-Risiko\nMISO-Setup")
shdr(ws3, r-1, amat_start_col + len(AMAT_MR)+1, "Empfehlung")
ws3.column_dimensions[get_column_letter(amat_start_col)].width   = 14
ws3.column_dimensions[get_column_letter(amat_start_col+1)].width = 14
ws3.column_dimensions[get_column_letter(amat_start_col+2)].width = 14
ws3.column_dimensions[get_column_letter(amat_start_col+3)].width = 14
ws3.column_dimensions[get_column_letter(amat_start_col+4)].width = 16
ws3.column_dimensions[get_column_letter(amat_start_col+5)].width = 28

QSPI_PROTO_CYCLES = 106   # from spec: 8+24+8+64+2

# Clock combinations to enumerate
# (f_in, pll_en, pll_n, f_sys, qspi_source, qspi_div)
# qspi_source: "=f_sys/div" or "=f_in direkt"
combos = []
for f_in in [10, 20, 25]:
    for pll_n in [1, 2, 4, 5]:
        f_sys = f_in * pll_n
        if f_sys > 60 or f_sys < 5:
            continue
        pll_en = "nein" if pll_n == 1 else "ja"
        for qdiv in [1, 2, 4]:
            f_qspi = f_sys / qdiv
            combos.append((f_in, pll_en, pll_n, f_sys, "f_sys / Div.", qdiv, f_qspi))

# Also add "QSPI = f_in direct" (crystal drives QSPI, separate path)
for f_in in [10, 25]:
    for pll_n in [1, 2, 5]:
        f_sys = f_in * pll_n
        if f_sys > 60 or f_sys < 5: continue
        pll_en = "nein" if pll_n == 1 else "ja"
        f_qspi = f_in   # crystal direct
        if f_qspi != f_sys:   # skip if same as div=1 already covered
            combos.append((f_in, pll_en, pll_n, f_sys, "f_in (direkt)", 1, f_qspi))

# Sort by f_sys desc, then f_qspi desc
combos.sort(key=lambda x: (-x[3], -x[6]))

# De-duplicate
seen = set()
unique_combos = []
for c in combos:
    key = (c[0], c[3], c[4], c[6])
    if key not in seen:
        seen.add(key); unique_combos.append(c)

# "Best" combinations for both flash devices:
# Criteria: f_qspi ≤ 80 MHz, MISO-timing safe, miss_penalty minimised, f_sys maximised
# We mark: ★★ best, ★ good
def timing_risk(f_sys, f_qspi, qspi_src):
    """MISO setup-time risk for synchronous QSPI controller."""
    if qspi_src == "f_in (direkt)":
        return "gering (async. Taktdomäne)"
    # tCO_max(W25Q128JV) ≈ 7 ns + ~3 ns PCB  → 10 ns total
    # Need: (1/f_qspi) - 10ns > setup_margin (5 ns min)
    period_ns = 1000.0 / f_qspi
    margin_ns = period_ns - 10          # generous: tCO=7ns + PCB=3ns
    if margin_ns >= 15:
        return "gering"
    elif margin_ns >= 8:
        return "moderat"
    elif margin_ns >= 2:
        return "erhöht – Layout prüfen"
    else:
        return "kritisch – nicht empfohlen"

def recommend(f_in, pll_en, pll_n, f_sys, qspi_src, f_qspi, t_miss_us):
    risk = timing_risk(f_sys, f_qspi, qspi_src)
    ok_flash = f_qspi <= 80
    if not ok_flash:
        return ("", "Außerhalb Flash-Spec")
    if "kritisch" in risk:
        return ("", "Timing-Risiko")
    # Penalize high dividers (= high miss penalty)
    miss_penalty = QSPI_PROTO_CYCLES * f_sys / f_qspi
    if f_sys >= 50 and f_qspi >= 25 and miss_penalty <= 212 and "erhöht" not in risk:
        if miss_penalty == 106 and f_sys == 50:
            return ("★★ BESTE OPTION", f"f_sys={f_sys} MHz, SCK={f_qspi} MHz, t_miss=106 Takte")
        if miss_penalty == 212 and f_sys == 50:
            return ("★★ EMPFOHLEN", f"f_sys={f_sys} MHz, f_QSPI={f_qspi} MHz, Timing sicher")
    pll_note = "kein PLL" if pll_en == "nein" else f"PLL×{pll_n}"
    if f_sys >= 25 and f_qspi >= 25 and miss_penalty <= 106 and "erhöht" not in risk:
        return ("★ GUT", f"{pll_note}, f_sys={f_sys} MHz, f_QSPI={f_qspi} MHz")
    if f_sys >= 25 and f_qspi >= 25:
        return ("◎ akzeptabel", "")
    return ("", "")

for idx, (f_in, pll_en, pll_n, f_sys, qspi_src, qdiv, f_qspi) in enumerate(unique_combos, start=1):
    t_qspi_ns   = 1000.0 / f_qspi if f_qspi > 0 else 0
    t_miss_us   = QSPI_PROTO_CYCLES / f_qspi if f_qspi > 0 else 0
    miss_pen    = QSPI_PROTO_CYCLES * f_sys / f_qspi if f_qspi > 0 else 0
    ok_w        = "✓" if f_qspi <= 80 else "✗  (zu schnell)"
    ok_m        = "✓" if f_qspi <= 80 else "✗  (zu schnell)"
    risk_str    = timing_risk(f_sys, f_qspi, qspi_src)
    rec_sym, rec_note = recommend(f_in, pll_en, pll_n, f_sys, qspi_src, f_qspi, t_miss_us)

    # Row background
    if "★★" in rec_sym:
        bg = FILL_BGRN
    elif "★" in rec_sym:
        bg = FILL_GREEN
    elif f_qspi > 80:
        bg = FILL_RED
    elif "kritisch" in risk_str:
        bg = FILL_ORANGE
    elif idx % 2 == 0:
        bg = FILL_GREY
    else:
        bg = FILL_WHITE

    row_vals = [
        idx, f_in,
        pll_en, (pll_n if pll_en == "ja" else "—"),
        f_sys, qspi_src, qdiv,
        f_qspi, f"{t_qspi_ns:.1f}", QSPI_PROTO_CYCLES,
        f"{t_miss_us:.2f}", f"{miss_pen:.0f}",
        ok_w, ok_m,
    ]
    for ci, v in enumerate(row_vals, start=1):
        ha = "center" if ci not in (6,) else "left"
        fnt = fbk(bold=("★" in rec_sym))
        s(ws3, r, ci, v, f=bg, fnt=fnt, ha=ha)
    # AMAT columns
    for ci2, mr in enumerate(AMAT_MR, start=amat_start_col):
        amat_val = 1 + mr * miss_pen
        num(ws3, r, ci2, round(amat_val, 1), nf="0.0", f=bg)
    # Timing risk
    risk_bg = (FILL_GREEN if "gering" in risk_str else
               FILL_YELLOW if "moderat" in risk_str else
               FILL_ORANGE if "erhöht" in risk_str else
               FILL_RED)
    s(ws3, r, amat_start_col+len(AMAT_MR), risk_str, f=risk_bg, fnt=fbk(), wrap=True)
    # Recommendation
    rec_bg = FILL_BGRN if "★★" in rec_sym else (FILL_GREEN if "★" in rec_sym else bg)
    rec_text = rec_sym + (f"\n{rec_note}" if rec_note else "")
    s(ws3, r, amat_start_col+len(AMAT_MR)+1, rec_text,
      f=rec_bg, fnt=fbk(bold=("★" in rec_sym)), wrap=True)
    ws3.row_dimensions[r].height = 18
    r+=1

r+=1
# Legend
legend_rows = [
    (FILL_BGRN,   "★★ BESTE OPTION / EMPFOHLEN  – optimales Verhältnis aus Performance und Timing-Sicherheit"),
    (FILL_GREEN,  "★ GUT  – geeignet für Referenzdesign"),
    (FILL_GREY,   "◎ akzeptabel – funktioniert, aber suboptimale Miss-Penalty oder Performance"),
    (FILL_ORANGE, "Erhöhtes Timing-Risiko – MISO-Setup muss im Place&Route verifiziert werden"),
    (FILL_RED,    "Nicht empfohlen – außerhalb Flash-Spec oder kritisches Timing"),
]
shdr(ws3, r, 1, "Farb-Legende", cs=amat_start_col+len(AMAT_MR)+1); r+=1
for bg, txt in legend_rows:
    s(ws3, r, 1, txt, f=bg, fnt=fbk(), cs=amat_start_col+len(AMAT_MR)+1)
    r+=1

# Freeze panes
ws3.freeze_panes = ws3.cell(row=7, column=2)

# ===========================================================================
# SHEET 4 – CPI_Analyse
# ===========================================================================
def build_cpi_sheet(wb):
    """
    Erstellt das Tabellenblatt 'CPI_Analyse'.

    Struktur
    --------
    Abschnitt 1  Einstellbare Parameter  (gelbe Zellen, manuell editierbar)
    Abschnitt 2  Abgeleitete Zwischengrößen  (Excel-Formeln, automatisch)
    Abschnitt 3  Design-Vergleich D1–D5  (Tabelle, alle Werte als Formeln)

    Sämtliche Zahlenergebnisse in Abschnitt 2 und 3 sind reine Excel-Formeln,
    die sich automatisch aktualisieren, sobald ein gelber Parameterwert geändert wird.

    Designs
    -------
    D1  RV_NoPipeline       – kein Pipeline, flaches SRAM (BPI-Bus), kein Cache
    D2  RV_NoPipelineCache  – kein Pipeline, I+D-Cache, SRAM-backed via AXI4
    D3  RV_PipelineCache    – 5-stufig, stall-only, KEIN Forwarding, I+D-Cache
    D4  Pipeline + SRAM     – hypothetisch: 5-stufig, direktes SRAM, kein Cache, kein Forwarding
    D5  Pipeline+Cache+Fwd  – hypothetisch: wie D3, aber mit EX→EX / MEM→EX Forwarding-Netz

    Zeilennummern der Parameterzellen (Spalte C)
    --------------------------------------------
    R_FLOAD=7   R_FSTORE=8  R_FBRANCH=9  R_FALU=10
    R_PTAKEN=13 R_BPEN=14
    R_HI=17     R_HD=18     R_MISS=19
    R_RAW1=23   R_RAW2=24   R_RAW3=25
    R_LU=29
    R_CPIRAW=33 R_CPIRFW=34 R_CPIBR=35 R_MISSID=36 R_MISSDD=37
    R_D1=41 R_D2=42 R_D3=43 R_D4=44 R_D5=45
    """
    ws = wb.create_sheet("CPI_Analyse")

    # ── Spaltenbreiten ────────────────────────────────────────────────────────
    # Cols 1-5 für Parameter-Block; cols 6-11 für den Design-Vergleich in Abschnitt 3
    for col, w in {1:4, 2:42, 3:14, 4:16, 5:62,
                   6:14, 7:14, 8:14, 9:14, 10:14, 11:42}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Lokale Füllfarben ─────────────────────────────────────────────────────
    FILL_PARAM = fill("FFF2CC")   # hellgelb  – editierbare Zellen
    FILL_CALC  = fill("EBF3FB")   # hellblau  – berechnete (read-only) Zellen
    FILL_D1    = fill("D9E1F2")   # blaugrau  – Design D1
    FILL_D2    = fill("E2EFD9")   # hellgrün  – Design D2
    FILL_D3    = fill("FCE4D6")   # lachsrot  – Design D3
    FILL_D4    = fill("FFF2CC")   # hellgelb  – Design D4 (hypothetisch)
    FILL_D5    = fill("C6EFCE")   # kräftiggrün – Design D5 (hypothetisch, beste Option)

    # ── Hilfsfunktion: Parameter-Zeile ────────────────────────────────────────
    def param_row(row, label, value, unit, note, fmt="0.00"):
        """
        Schreibt eine editierbare Parameterzeile (Wert-Zelle gelb hinterlegt).
        Spalte B: Bezeichnung, C: Wert (editierbar), D: Einheit, E: Beschreibung.
        """
        s(ws, row, 2, label, fnt=fbk(bold=True), wrap=True)
        c = ws.cell(row=row, column=3)
        c.value         = value
        c.fill          = FILL_PARAM
        c.font          = fbk(bold=True, color="7F4B00")
        c.alignment     = Alignment(horizontal="center", vertical="center")
        c.border        = THIN
        c.number_format = fmt
        s(ws, row, 4, unit, fnt=fbk(bold=False))
        s(ws, row, 5, note, fnt=fbk(bold=False), wrap=True)
        ws.row_dimensions[row].height = 18

    # ── Hilfsfunktion: Berechnete Zeile (Abschnitt 2) ─────────────────────────
    def calc_row(row, label, formula, unit, note):
        """
        Schreibt eine berechnete Zeile (Formel in Spalte C, hellblau hinterlegt).
        Alle Formeln verweisen auf absolute Zellreferenzen aus Abschnitt 1.
        """
        s(ws, row, 2, label, fnt=fbk(bold=True))
        c = ws.cell(row=row, column=3)
        c.value         = formula
        c.fill          = FILL_CALC
        c.font          = fbk(bold=False, color="1F4E79")
        c.alignment     = Alignment(horizontal="center", vertical="center")
        c.border        = THIN
        c.number_format = "0.000"
        s(ws, row, 4, unit, fnt=fbk())
        s(ws, row, 5, note, fnt=fbk(), wrap=True)
        ws.row_dimensions[row].height = 18

    # ── Hilfsfunktion: Design-Vergleichszeile (Abschnitt 3) ───────────────────
    # ref_d1 / ref_d4 werden nach den Zeilennummern-Konstanten definiert.
    def design_row(row, did, name, f_arch, f_hazard, f_miss, note, bg, ref_d1, ref_d4):
        """
        Schreibt eine vollständige Design-Zeile in den Vergleichs-Block.

        Spalten
        -------
        B  Design-ID (z.B. "D1")
        C  Beschreibung
        D  CPI_arch   – strukturelle Zyklen aus FSM / Pipeline-Aufbau
        E  CPI_hazard – Stall-Zyklen durch RAW- und Control-Hazards
        F  CPI_miss   – Stall-Zyklen durch Cache-Misses
        G  CPI_total  = D+E+F  (Excel-Formel)
        H  IPC        = 1/G    (Excel-Formel)
        I  Speedup vs D1 = CPI_D1/G  (> 1 → schneller als D1)
        J  Speedup vs D4 = CPI_D4/G  (> 1 → schneller als hypothetisches SRAM-Design)
        K  Anmerkungen
        """
        f_total  = f"=D{row}+E{row}+F{row}"
        f_ipc    = f"=1/G{row}"
        f_spd_d1 = f"={ref_d1}/G{row}"
        f_spd_d4 = f"={ref_d4}/G{row}"

        data = [
            (2,  did,      "@",    FILL_DARK,  fwh(bold=True, sz=11), "center", False),
            (3,  name,     "@",    bg,         fbk(bold=False),        "left",  True),
            (4,  f_arch,   "0.000",bg,         fbk(color="1F4E79"),    "center",False),
            (5,  f_hazard, "0.000",bg,         fbk(color="1F4E79"),    "center",False),
            (6,  f_miss,   "0.000",bg,         fbk(color="1F4E79"),    "center",False),
            (7,  f_total,  "0.000",bg,         fbk(bold=True),         "center",False),
            (8,  f_ipc,    "0.000",bg,         fbk(),                  "center",False),
            (9,  f_spd_d1, "0.00", bg,         fbk(),                  "center",False),
            (10, f_spd_d4, "0.00", bg,         fbk(),                  "center",False),
            (11, note,     "@",    bg,         fbk(bold=False),        "left",  True),
        ]
        for ci, val, nf, fill_, fnt, ha, wrap_ in data:
            c = ws.cell(row=row, column=ci)
            c.value         = val
            c.fill          = fill_
            c.font          = fnt
            c.number_format = nf
            c.alignment     = Alignment(horizontal=ha, vertical="center",
                                        wrap_text=wrap_)
            c.border        = THIN
        ws.row_dimensions[row].height = 22

    # =========================================================================
    # Zeilennummern-Konstanten
    # Alle Formeln in Abschnitt 2 und 3 referenzieren diese absoluten Adressen.
    # Die assert-Anweisungen unten sichern, dass der Zähler r exakt passt.
    # =========================================================================
    # Abschnitt 1 – Parameter
    R_FLOAD   = 7    # f_load   (editierbar)
    R_FSTORE  = 8    # f_store  (editierbar)
    R_FBRANCH = 9    # f_branch (editierbar)
    R_FALU    = 10   # f_alu    (Formel: 1−load−store−branch)
    R_PTAKEN  = 13   # p_taken  (editierbar)
    R_BPEN    = 14   # cyc_branch (editierbar)
    R_HI      = 17   # h_I      (editierbar)
    R_HD      = 18   # h_D      (editierbar)
    R_MISS    = 19   # m_cyc    (editierbar)
    R_RAW1    = 23   # p_raw1   (editierbar)
    R_RAW2    = 24   # p_raw2   (editierbar)
    R_RAW3    = 25   # p_raw3   (editierbar)
    R_LU      = 29   # p_lu     (editierbar)
    # Abschnitt 2 – Zwischengrößen
    R_CPIRAW  = 33   # CPI_raw   (kein Forwarding)
    R_CPIRFW  = 34   # CPI_raw_fw (mit Forwarding)
    R_CPIBR   = 35   # CPI_branch
    R_MISSID  = 36   # CPI_miss_I
    R_MISSDD  = 37   # CPI_miss_D
    # Abschnitt 3 – Design-Zeilen (Spalte G = col 7 = CPI_total)
    R_D1 = 41
    R_D2 = 42
    R_D3 = 43
    R_D4 = 44
    R_D5 = 45
    COL_TOT = 7   # Spalte G: CPI_total

    # Absolute Referenz-Strings für Speedup-Formeln
    ref_d1 = f"${get_column_letter(COL_TOT)}${R_D1}"   # z.B. "$G$41"
    ref_d4 = f"${get_column_letter(COL_TOT)}${R_D4}"   # z.B. "$G$44"

    # =========================================================================
    # TITEL
    # =========================================================================
    r = 1
    s(ws, r, 1, "CPI-Analyse  —  RWU-RV64I Familie  (5 Design-Varianten)",
      FILL_DARK, fwh(sz=13), ha="center", cs=11)
    ws.row_dimensions[r].height = 28; r += 1

    s(ws, r, 1,
      "Grundlage: RTL-Analyse  as_cpux.sv (RV_NoPipeline / RV_NoPipelineCache)  und  "
      "cc_cpupipe.sv (RV_PipelineCache)  |  Speicher: X-FAB synchrones SRAM  |  "
      "Alle gelben Zellen in Abschnitt 1 sind editierbar",
      FILL_MED, fwh(sz=9, bold=False), ha="center", cs=11)
    r += 2   # r = 4

    # =========================================================================
    # ABSCHNITT 1 – PARAMETER
    # =========================================================================
    shdr(ws, r, 1,
         "ABSCHNITT 1  —  Einstellbare Parameter  "
         "( gelbe Zellen editieren → Abschnitt 2 und 3 aktualisieren sich automatisch )",
         cs=11)
    r += 1   # r = 5

    # ── Block A: Code-Mix ─────────────────────────────────────────────────────
    sh2(ws, r, 1, "A  Code-Mix  (typischer RISC-V Instruction-Mix)", cs=11); r += 1  # r = 6

    s(ws, r, 2, "Randbedingung:  f_load + f_store + f_branch + f_alu = 1",
      fnt=fbk(bold=False, color="666666"), cs=4)
    r += 1   # r = 7

    assert r == R_FLOAD
    param_row(r, "f_load     Anteil Load-Instruktionen", 0.25, "–",
        "LB/LH/LW/LD/LBU/LHU  –  typisch 25 % bei embedded RV64I-Code (Harris & Patterson)")
    r += 1   # r = 8

    assert r == R_FSTORE
    param_row(r, "f_store    Anteil Store-Instruktionen", 0.10, "–",
        "SB/SH/SW/SD  –  typisch 10 %")
    r += 1   # r = 9

    assert r == R_FBRANCH
    param_row(r, "f_branch   Anteil Branch-Instruktionen", 0.15, "–",
        "BEQ/BNE/BLT/BGE/BLTU/BGEU + JAL (für Loops)  –  typisch 15 %")
    r += 1   # r = 10

    assert r == R_FALU
    s(ws, r, 2, "f_alu      Anteil ALU / sonstige  (automatisch)", fnt=fbk(bold=True))
    c = ws.cell(row=r, column=3)
    c.value         = f"=1-$C${R_FLOAD}-$C${R_FSTORE}-$C${R_FBRANCH}"
    c.fill          = FILL_CALC
    c.font          = fbk(bold=False, color="1F4E79")
    c.alignment     = Alignment(horizontal="center", vertical="center")
    c.border        = THIN
    c.number_format = "0.00"
    s(ws, r, 4, "–")
    s(ws, r, 5, "= 1 − f_load − f_store − f_branch  "
      "(ALU R/I-Type, LUI, AUIPC, CSR, JALR ohne Abhängigkeit)", fnt=fbk())
    ws.row_dimensions[r].height = 18
    r += 2   # r = 12

    # ── Block B: Branch-Verhalten ─────────────────────────────────────────────
    sh2(ws, r, 1, "B  Sprung-Verhalten", cs=11); r += 1  # r = 13

    assert r == R_PTAKEN
    param_row(r, "p_taken    Branch taken-Rate", 0.60, "–",
        "Anteil tatsächlich ausgeführter Sprünge.  "
        "Typisch ≈ 60 % (Schleifen dominieren den Programmfluss).  "
        "Bei p_taken = 0 kein Flush-Overhead (alle Sprünge not-taken).")
    r += 1   # r = 14

    assert r == R_BPEN
    param_row(r, "cyc_branch Branch-Flush-Penalty  [Takte]", 2, "Takte",
        "Fest durch RTL cc_cpupipe.sv:  flush_s = ex_branch_taken_s  →  "
        "IF- und ID-Stufe werden mit NOP überschrieben  →  2 verlorene Stufen.", fmt="0")
    r += 2   # r = 16

    # ── Block C: Speicher-Parameter ───────────────────────────────────────────
    sh2(ws, r, 1, "C  Speicher-Parameter  (X-FAB synchrones SRAM)", cs=11); r += 1  # r = 17

    assert r == R_HI
    param_row(r, "h_I        I-Cache Hit-Rate", 0.98, "–",
        "Trefferrate für Instruction-Fetches.  "
        "0.98 = typisch für kleine embedded Loops.  "
        "D1 und D4 ignorieren diesen Parameter  (kein Cache → effektiv h_I = 1.0).",
        fmt="0.000")
    r += 1   # r = 18

    assert r == R_HD
    param_row(r, "h_D        D-Cache Hit-Rate  (Loads)", 0.98, "–",
        "Trefferrate für Data-Loads.  "
        "Store-Misses = 0: Stores gehen direkt in den Scratchpad-SRAM (kein Evict).  "
        "D1 und D4 ignorieren diesen Parameter.",
        fmt="0.000")
    r += 1   # r = 19

    assert r == R_MISS
    param_row(r, "m_cyc      Miss-Penalty  [Systemtakte]", 6, "Takte",
        "Zyklen für Cache-Refill bei einem Miss.  "
        "On-Chip SRAM hinter AXI4 (32B-Zeile, 64-bit-Bus):  "
        "≈ 1(AR) + 1(SRAM-Adr.) + 4(R-Beats à 8 Byte) = 6 Takte.  "
        "Flash-QSPI: 106×(f_sys/f_QSPI_SCK)  →  siehe Sheet 'AMAT_Szenarien'.",
        fmt="0")
    r += 2   # r = 21

    # ── Block D: RAW-Hazards ohne Forwarding ──────────────────────────────────
    sh2(ws, r, 1, "D  RAW Data-Hazards  –  OHNE Forwarding  (gilt für D3, D4)", cs=11)
    r += 1   # r = 22

    s(ws, r, 2,
      "cc_cpupipe.sv besitzt kein Forwarding-Netz.  "
      "Jede RAW-Abhängigkeit erzwingt Stall-Zyklen, bis der Produzent die WB-Stufe abgeschlossen hat.",
      fnt=fbk(bold=False, color="444444"), wrap=True, cs=4)
    ws.row_dimensions[r].height = 24
    r += 1   # r = 23

    assert r == R_RAW1
    param_row(r, "p_raw1     RAW 1-back Häufigkeit", 0.25, "–",
        "Instr. N liest ein Register, das Instr. N−1 schreibt  →  3 Stall-Zyklen.  "
        "Produzent ist in EX, Konsument bleibt 3 Takte in ID stehen (EX→MEM→WB).  "
        "Typisch ≈ 25 % (Harris & Patterson, Kap. 7).")
    r += 1   # r = 24

    assert r == R_RAW2
    param_row(r, "p_raw2     RAW 2-back Häufigkeit", 0.08, "–",
        "Instr. N liest ein Register, das Instr. N−2 schreibt  →  2 Stall-Zyklen.  "
        "Typisch ≈ 8 %.")
    r += 1   # r = 25

    assert r == R_RAW3
    param_row(r, "p_raw3     RAW 3-back Häufigkeit", 0.04, "–",
        "Instr. N liest ein Register, das Instr. N−3 schreibt  →  1 Stall-Zyklus.  "
        "Typisch ≈ 4 %.")
    r += 2   # r = 27

    # ── Block E: RAW mit Forwarding ───────────────────────────────────────────
    sh2(ws, r, 1, "E  RAW Data-Hazards  –  MIT Forwarding  (gilt nur für D5, hypothetisch)", cs=11)
    r += 1   # r = 28

    s(ws, r, 2,
      "Mit vollständigem EX→EX und MEM→EX Forwarding-Netz.  "
      "Einziger verbleibender Stall: Load-Use-Hazard  "
      "(Load-Ergebnis erst nach MEM verfügbar, Konsument braucht es sofort in EX  →  1 Stall-Zyklus).",
      fnt=fbk(bold=False, color="444444"), wrap=True, cs=4)
    ws.row_dimensions[r].height = 28
    r += 1   # r = 29

    assert r == R_LU
    param_row(r, "p_lu       Load-Use Häufigkeit", 0.05, "–",
        "Anteil aller Instruktionen, die ein Load-Ergebnis unmittelbar (1-back) lesen.  "
        "Mit Forwarding der einzige verbleibende 1-Stall-Fall.  "
        "Typisch ≈ 5 % (≈ 20 % aller Loads haben eine sofortige Abhängigkeit).")
    r += 2   # r = 31

    # =========================================================================
    # ABSCHNITT 2 – ZWISCHENGROSSEN
    # =========================================================================
    shdr(ws, r, 1,
         "ABSCHNITT 2  —  Abgeleitete Zwischengrößen  "
         "(Excel-Formeln — aktualisieren sich automatisch bei Parameteränderung)",
         cs=11)
    r += 1   # r = 32

    s(ws, r, 2, "Größe",                   f=FILL_LIGHT, fnt=fbk(bold=True))
    s(ws, r, 3, "Wert",                    f=FILL_LIGHT, fnt=fbk(bold=True), ha="center")
    s(ws, r, 4, "Einheit",                 f=FILL_LIGHT, fnt=fbk(bold=True))
    s(ws, r, 5, "Formel und Erläuterung",  f=FILL_LIGHT, fnt=fbk(bold=True), cs=7)
    r += 1   # r = 33

    assert r == R_CPIRAW
    calc_row(r,
        "CPI_raw      RAW-Stall-Anteil  (kein Forwarding)",
        f"=$C${R_RAW1}*3 + $C${R_RAW2}*2 + $C${R_RAW3}*1",
        "Takte / Instr.",
        f"p_raw1×3 + p_raw2×2 + p_raw3×1  —  Stall-Tiefe: 3 wenn Produzent in EX, "
        f"2 wenn in MEM, 1 wenn in WB.  Kein Forwarding: Konsument wartet in ID, "
        f"bis Produzent WB abgeschlossen hat.")
    r += 1   # r = 34

    assert r == R_CPIRFW
    calc_row(r,
        "CPI_raw_fw   RAW-Stall-Anteil  (mit Forwarding, D5)",
        f"=$C${R_LU}*1",
        "Takte / Instr.",
        f"p_lu × 1  —  Mit EX→EX/MEM→EX-Forwarding bleibt nur der Load-Use-Fall "
        f"(1 Stall-Zyklus), weil das Load-Ergebnis erst nach der MEM-Stufe vorliegt, "
        f"nicht schon nach EX.")
    r += 1   # r = 35

    assert r == R_CPIBR
    calc_row(r,
        "CPI_branch   Branch-Penalty-Anteil",
        f"=$C${R_FBRANCH}*$C${R_PTAKEN}*$C${R_BPEN}",
        "Takte / Instr.",
        f"f_branch × p_taken × cyc_branch  —  Branch wird in EX aufgelöst "
        f"(cc_cpupipe.sv: flush_s = ex_branch_taken_s); 2 Stufen (IF, ID) werden "
        f"mit NOP überschrieben.  Gilt für D3, D4, D5.")
    r += 1   # r = 36

    assert r == R_MISSID
    calc_row(r,
        "CPI_miss_I   I-Cache-Miss-Anteil  (pro Instruktion)",
        f"=(1-$C${R_HI})*$C${R_MISS}",
        "Takte / Instr.",
        f"(1−h_I) × m_cyc  —  Bei jedem Instruction-Fetch-Miss friert die gesamte "
        f"Pipeline ein bis ic_rvalid = 1 (icache_stall_s in cc_cpupipe.sv).  "
        f"Gilt für D2, D3, D5.")
    r += 1   # r = 37

    assert r == R_MISSDD
    calc_row(r,
        "CPI_miss_D   D-Cache-Miss-Anteil  (pro Instruktion)",
        f"=$C${R_FLOAD}*(1-$C${R_HD})*$C${R_MISS}",
        "Takte / Instr.",
        f"f_load × (1−h_D) × m_cyc  —  D-Cache-Miss tritt nur bei Loads auf; "
        f"Stores gehen direkt in den SRAM (Scratchpad-Bypass, kein Evict).  "
        f"Gilt für D2, D3, D5.")
    r += 2   # r = 39

    # =========================================================================
    # ABSCHNITT 3 – DESIGN-VERGLEICH
    # =========================================================================
    shdr(ws, r, 1,
         "ABSCHNITT 3  —  Design-Vergleich  D1–D5  "
         "( Speedup > 1.0 = schneller als Referenz )",
         cs=11)
    r += 1   # r = 40

    tbl_hdrs = [
        "Design",
        "Beschreibung",
        "CPI  Architektur\n(FSM / Pipeline)",
        "CPI  Hazard-\nStalls",
        "CPI  Cache-\nMiss",
        "CPI  gesamt",
        "IPC",
        "Speedup\nvs. D1",
        "Speedup\nvs. D4",
        "Anmerkungen",
    ]
    for ci, h in enumerate(tbl_hdrs, start=2):
        shdr(ws, r, ci, h)
        ws.cell(row=r, column=ci).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 38
    r += 1   # r = 41

    # ── D1: RV_NoPipeline ─────────────────────────────────────────────────────
    # Strukturelle CPI aus der 4-Zustand-FSM (as_cpux.sv):
    #   Nicht-Load: FETCH0(1) + FETCH1(1) + EXEC(1)      = 3 Takte
    #   Load:       FETCH0(1) + FETCH1(1) + EXEC(1) + EXECLD(1) = 4 Takte
    #   Gewichtet:  (1−f_load)×3 + f_load×4 = 3 + f_load
    # Kein Pipeline → keine Hazard-Stalls; kein Cache → keine Miss-Stalls.
    assert r == R_D1
    design_row(r, "D1",
        "RV_NoPipeline  —  kein Pipeline, flaches SRAM (BPI-Bus), kein Cache",
        f"=3+$C${R_FLOAD}",   # 3 + f_load (gewichtetes Mittel über alle Instr.-Klassen)
        "=0",                  # kein Pipeline → keine RAW- oder Branch-Stalls
        "=0",                  # kein Cache → keine Miss-Stalls
        "4-Zustand FSM F0→F1→EX→[ELD]→F0.  "
        "Deterministisch: exakt 3 Takte (nicht-Load) bzw. 4 Takte (Load).  "
        "Speedup vs. D1 = 1.00 (Referenz-Design).",
        FILL_D1, ref_d1, ref_d4)
    r += 1   # r = 42

    # ── D2: RV_NoPipelineCache ────────────────────────────────────────────────
    # Gleiche FSM wie D1 (identischer struktureller CPI).
    # FETCH1 und EXECLD können N Takte stallieren, bis ic_rvalid / dc_rvalid = 1.
    # CPI_miss = (1−h_I)·m  +  f_load·(1−h_D)·m
    assert r == R_D2
    design_row(r, "D2",
        "RV_NoPipelineCache  —  kein Pipeline, I+D-Cache, SRAM-backed via AXI4",
        f"=3+$C${R_FLOAD}",                            # gleiche FSM-Struktur wie D1
        "=0",                                           # kein Pipeline → kein Hazard-Stall
        f"=$C${R_MISSID}+$C${R_MISSDD}",               # Cache-Miss-Anteil
        "Identische FSM wie D1, aber FETCH1 und EXECLD warten auf ic_rvalid/dc_rvalid.  "
        "CPI_miss = CPI_miss_I + CPI_miss_D  (aus Abschnitt 2).  "
        "Bei h_I = h_D = 1.0 identisch mit D1.",
        FILL_D2, ref_d1, ref_d4)
    r += 1   # r = 43

    # ── D3: RV_PipelineCache ──────────────────────────────────────────────────
    # 5-stufige Pipeline IF/ID/EX/MEM/WB, stall-only, KEIN Forwarding.
    # CPI_arch = 1 (idealer Pipeline-Durchsatz).
    # CPI_hazard = CPI_raw (RAW-Stalls, kein Forwarding) + CPI_branch (Control-Stalls).
    # CPI_miss = CPI_miss_I + CPI_miss_D (globaler Pipeline-Freeze bei Cache-Miss).
    assert r == R_D3
    design_row(r, "D3",
        "RV_PipelineCache  —  5-stufig, stall-only, KEIN Forwarding, I+D-Cache",
        "=1",                                           # idealer Pipeline-Durchsatz
        f"=$C${R_CPIRAW}+$C${R_CPIBR}",                # RAW (kein Fwd) + Branch-Penalty
        f"=$C${R_MISSID}+$C${R_MISSDD}",               # Cache-Miss-Penalty
        "5-stufig IF/ID/EX/MEM/WB (cc_cpupipe.sv).  Kein Forwarding: "
        "RAW 1/2/3-back → 3/2/1 Stall-Zyklen in ID.  "
        "Branch in EX aufgelöst → 2-Stufen-Flush.  "
        "Cache-Miss → globaler Freeze (icache_stall_s | dcache_stall_s).",
        FILL_D3, ref_d1, ref_d4)
    r += 1   # r = 44

    # ── D4: Hypothetisch Pipeline + direktes SRAM ─────────────────────────────
    # Gleiche 5-stufige Architektur wie D3, aber I-SRAM und D-SRAM direkt
    # angebunden (1-Takt-Latenz, immer, kein AXI4-Overhead, keine Misses).
    # Äquivalent zu D3 mit h_I = h_D = 1.0.
    assert r == R_D4
    design_row(r, "D4",
        "Pipeline + SRAM (hypothetisch)  —  5-stufig, direktes SRAM, kein Cache, kein Forwarding",
        "=1",                                           # gleicher Pipeline-Aufbau wie D3
        f"=$C${R_CPIRAW}+$C${R_CPIBR}",                # gleiche Hazard-Stalls wie D3
        "=0",                                           # kein Cache → keine Miss-Stalls
        "Wie D3, aber I-SRAM und D-SRAM direkt per SRAM-Interface angebunden.  "
        "Kein AXI4, keine Cache-Misses, deterministisches Timing.  "
        "Entspricht D3 mit h_I = h_D = 1.0.  "
        "Einschränkung: Programm- und Datengröße durch SRAM-Kapazität begrenzt.",
        FILL_D4, ref_d1, ref_d4)
    r += 1   # r = 45

    # ── D5: Hypothetisch Pipeline + Cache + Forwarding ────────────────────────
    # Wie D3, aber mit vollständigem Forwarding-Netz (EX→EX, MEM→EX).
    # Verbleibende Stalls: Load-Use (1 Zyklus) + Branch (unverändert).
    # Entspricht einer typischen 5-stufigen In-Order-Pipeline (MIPS R3000-Klasse).
    assert r == R_D5
    design_row(r, "D5",
        "Pipeline + Cache + Forwarding (hypothetisch)  —  5-stufig, EX→EX / MEM→EX Forwarding",
        "=1",                                           # gleicher Pipeline-Aufbau
        f"=$C${R_CPIRFW}+$C${R_CPIBR}",                # Load-Use (1 Stall) + Branch-Penalty
        f"=$C${R_MISSID}+$C${R_MISSDD}",               # gleiche Cache-Miss-Penalty wie D3
        "Wie D3, aber mit EX→EX/MEM→EX Forwarding-Netz.  "
        "Einziger verbleibender Stall: Load-Use (1 Zyklus).  "
        "Typische In-Order-Pipeline (MIPS R3000-Klasse).  "
        "Implementierungsaufwand: ~50 LUT zusätzlich in cc_cpupipe.sv; "
        "kritischer Pfad verlängert sich durch Forwarding-Mux leicht.",
        FILL_D5, ref_d1, ref_d4)
    r += 2

    # ── Farblegende ───────────────────────────────────────────────────────────
    shdr(ws, r, 1, "Farb-Legende", cs=11); r += 1
    for bg, txt in [
        (FILL_D1, "D1  RV_NoPipeline           – Referenz-Design  (CPI_total = Referenz für Speedup-Spalte I)"),
        (FILL_D2, "D2  RV_NoPipelineCache       – Cache-Erweiterung ohne Pipeline"),
        (FILL_D3, "D3  RV_PipelineCache         – implementierte Pipeline  (cc_cpupipe.sv, kein Forwarding)"),
        (FILL_D4, "D4  Pipeline + SRAM          – hypothetisch, kein Cache, kein Forwarding  "
                  "(= Referenz für Speedup-Spalte J: zeigt Cache-Miss-Overhead von D3/D5)"),
        (FILL_D5, "D5  Pipeline+Cache+Fwd       – hypothetisch, optimale In-Order-Pipeline"),
        (FILL_PARAM, "Gelb  – editierbare Parameterzelle (Abschnitt 1)"),
        (FILL_CALC,  "Hellblau  – berechnete Zelle (Formel, nicht editieren)"),
    ]:
        s(ws, r, 1, txt, f=bg, fnt=fbk(), cs=11); r += 1

    r += 1

    # ── Annahmen und Einschränkungen ──────────────────────────────────────────
    shdr(ws, r, 1, "Annahmen und Einschränkungen", cs=11); r += 1
    notes = [
        ("Speicher-Modell",
         "Alle Designs: X-FAB synchrones SRAM, 1-Takt Read-Latenz "
         "(Adresse an steigender Flanke N → Daten gültig vor steigender Flanke N+1). "
         "D1/D4 greifen direkt per BPI/SRAM-Interface zu. "
         "D2/D3/D5 nutzen I+D-Cache mit AXI4-Refill-Pfad zum Haupt-SRAM."),
        ("Code-Mix",
         "Typische Werte aus Harris & Patterson 'Digital Design and Computer Architecture' "
         "und SPEC-CPU-Messungen für embedded RISC-V. "
         "Für konkrete RWU-Testprogramme (UART, GPIO) empfiehlt sich eine Profiling-basierte Anpassung "
         "von f_load, f_store, f_branch."),
        ("RAW-Häufigkeiten (D3/D4)",
         "p_raw1/2/3 gelten für typischen C-Code kompiliert mit gcc -O1. "
         "Bei -O0 (kein Instruction Scheduling) steigen die Werte erheblich. "
         "Bei -O2/-O3 (Reorder Buffer) sinken sie signifikant. "
         "Assemblerprogramme (UART-Test) liegen je nach Schreibstil dazwischen."),
        ("Branch-Modell",
         "Kein Branch Predictor in allen fünf Designs. "
         "Penalty gilt ausschließlich für genommene Sprünge (taken). "
         "Not-taken Branches verursachen in der Pipeline keinen Flush-Overhead. "
         "JAL (immer taken) ist in f_branch enthalten und zählt immer zur Penalty."),
        ("Stores",
         "D1/D2: Store schreibt in EXEC_ST (synchrones Schreiben, kein EXECLD) → CPI_store = 3. "
         "D3/D4/D5: Store wird in der MEM-Stufe abgeschlossen (1 Takt). "
         "Kein Miss-Stall für Stores in allen Designs (Scratchpad-/SRAM-Bypass, kein Evict)."),
        ("Forwarding D5",
         "Das hypothetische Forwarding-Netz leitet ALU-Ergebnisse aus EX und MEM "
         "direkt in den EX-Eingangs-Mux zurück (zwei neue Mux-Stufen in cc_cpupipe.sv). "
         "Kritischer Pfad: ALU → Forwarding-Mux → nächste ALU. "
         "Implementierungsaufwand: ≈ 50 LUT (Schätzung für XO035, 4-LUT-Technologie)."),
        ("Speedup-Interpretation",
         "Speedup = CPI_Referenz / CPI_Design.  Wert > 1.0 bedeutet schneller als die Referenz. "
         "Spalte I (vs. D1): absoluter Gewinn der Pipeline über das sequentielle Design. "
         "Spalte J (vs. D4): zeigt den Cache-Miss-Overhead von D2/D3/D5 gegenüber "
         "dem hypothetischen fehlerfreien SRAM-Design D4."),
    ]
    for i, (title, body) in enumerate(notes):
        bg = FILL_GREY if i % 2 == 0 else FILL_WHITE
        s(ws, r, 2, title, f=FILL_LIGHT, fnt=fbk(bold=True), cs=2)
        s(ws, r, 4, body,  f=bg, fnt=fbk(bold=False), cs=8, wrap=True)
        ws.row_dimensions[r].height = 44; r += 1

    ws.freeze_panes = "B5"


build_cpi_sheet(wb)

# ===========================================================================
# Save
# ===========================================================================
wb.save(OUT)
print(f"Saved: {OUT}")
